
from copy import deepcopy
import requests
import os
import bs4
from openpyxl import load_workbook
import pandas as pd
from ..helpers.db_funcs import get_ep_id_by_number, get_season_id_by_number_type
from ..helpers.extract_helpers import search_for_new_seasons
import glob
import re
import numpy as np


DOCS_URL_TEMPLATE = 'https://docs.google.com/spreadsheets/d/{id}/export?format=xlsx&id={id}'
SURVIVOR_SOURCE = 'https://www.truedorktimes.com/survivor/boxscores/data.htm'


def create_data_dict(subset=None):
    ret_dict = {}
    sp = bs4.BeautifulSoup(requests.get(SURVIVOR_SOURCE).content)
    cast_elements = sp.find_all('ul', attrs={'class': 'cast'})

    for e in cast_elements:
        attrs = e.find('a').attrs

        try:
            if 'spreadsheet' in attrs['href']:
                v = attrs['href'][:-1].split('/')[-1]
                k = str(e.text.lower())
                for p in '- ':
                    k = k.replace(p, '_')
                for p in ':.-,':
                    k = k.replace(p, '')
                k = k.replace('\n', '')[1:]

                if subset:
                    if k.split('_')[0] not in subset:
                        continue

                ret_dict[k] = v

            else:
                pass
        except KeyError:
            pass

    return ret_dict


def save_survivor_excel(sheets_id, readable_name, dest_folder='../data/raw'):
    url = DOCS_URL_TEMPLATE.format(**dict(id=sheets_id))
    f_name = '{readable_name}.xlsx'.format(readable_name=readable_name)

    req = requests.get(url)

    with open(os.path.join(dest_folder, f_name), 'wb') as f:
        f.write(req.content)
    req.close()


def pull_and_save_excels(data_dict=None, subset=None, dest_folder='../data/raw'):
    if not data_dict:
        data_dict = create_data_dict(subset=subset)
    for k, v in data_dict.items():
        save_survivor_excel(v, k, dest_folder=dest_folder)


# Above is for the actual excels...


def empty_cond(ws, cell, *args, **kwargs):
    return not cell.value


def vertical_cond_tc(ws, cell, *args, **kwargs):

    return empty_cond(ws, cell) or (cell.value == 'wanda')


def any_cond(ws, cell):
    return False


def rc_horizontal_cond(ws, cell, col_names, *args, **kwargs):
    above = ws.cell(row=cell.row - 1, column=cell.column).value
    if isinstance(above, str):
        ic_bool = ('IC' in above) or (
            'Immunity challenge' in above) or ('RC' in above)
        ic_bool = ic_bool and (len(col_names) != 0)
    else:
        ic_bool = False
    return (not cell.value) or (ic_bool)


def ep_horizontal_cond(ws, cell, col_names, nblanks=5, *args, **kwargs):
    add_numbers = [x for x in range(1, nblanks + 1)]
    right_two = [not ws.cell(
        row=cell.row, column=cell.column + i).value for i in add_numbers]
    return all(right_two) and not cell.value


def normal_extract_values(ws, row, column_start, width, *args, **kwargs):
    return pd.Series([ws.cell(row=row, column=column_start + i + 1).value for i in range(width)])


def vote_extract_values(ws, row, column_start, width, col_names, *args, **kwargs):
    values = normal_extract_values(ws, row, column_start, width)
    values = pd.Series([c for i, c in enumerate(
        col_names) if pd.notnull(values[i])])
    return values if len(values) > 0 else pd.Series([None])


def identity_pp(df, col_names, *args, **kwargs):
    df.columns = col_names

    df = df.loc[:, ~df.columns.duplicated()]
    return df


def ep_pp(df, col_names, *args, **kwargs):
    df = identity_pp(df, col_names, *args, **kwargs)
    df = df[df.columns[~df.columns.isna()]]
    return df


def vote_pp(df, col_names, *args, **kwargs):
    if df.shape[1] > 1:
        df = pd.DataFrame(pd.concat([df[col] for col in df]))

    df.columns = ['voted_for']
    df['vote_counted'] = ~df['voted_for'].isna()
    df = df.loc[:, ~df.columns.duplicated()]
    return df


def extract_subtable(ws, start_row, start_column, index_column=None, horizontal_condition=None,
                     vertical_condition=None, extract_values=None, postprocess=None):

    if not horizontal_condition:
        horizontal_condition = empty_cond

    if not vertical_condition:
        vertical_condition = empty_cond

    if not extract_values:
        extract_values = normal_extract_values

    if not postprocess:
        postprocess = identity_pp
    row = start_row
    col = start_column

    col_names = []
    rows = []
    idx = []

    while True:
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if horizontal_condition(ws, cell, col_names):
            break
        else:
            col_names.append(v)
            col += 1

    n_voted_against = len(col_names)
    col -= (n_voted_against + 1)
    row += 1

    while True:
        idx_cell = ws.cell(
            row=row, column=index_column if index_column else col)
        if vertical_condition(ws, idx_cell):
            break
        else:
            values = extract_values(ws, row, col, n_voted_against, col_names)
            rows.append(values)
            idx.append(idx_cell.value)

            row += 1
    df = pd.DataFrame(rows)
    df.index = idx
    df.index.name = 'contestant'

    df = postprocess(df, col_names)

    df.reset_index(inplace=True)

    return df


def extract_rc_challenge(ws, c_row, c_column):
    return extract_subtable(ws, c_row + 1, c_column, horizontal_condition=rc_horizontal_cond, index_column=1)


def extract_ic_challenge(ws, c_row, c_column):
    return extract_subtable(ws, c_row + 1, c_column, horizontal_condition=rc_horizontal_cond, index_column=1)


def extract_tc(ws, c_row, c_column):
    return extract_subtable(ws, c_row + 1, c_column, vertical_condition=vertical_cond_tc, extract_values=vote_extract_values, postprocess=vote_pp, index_column=1)


def extract_ep(ws, c_row, c_column):
    return extract_subtable(ws, c_row, c_column, index_column=1, horizontal_condition=ep_horizontal_cond, postprocess=ep_pp)


def append_tc_data(df, ws, cell):
    v = cell.value
    try:
        total_players = int(re.search('F(\d+)', v).group(1))
    except AttributeError:
        if 'No' in v:
            return pd.DataFrame()
        elif 'Morgan' in v:
            total_players = 10
        elif 'Drake' in v:
            total_players = 9
        elif 'Mokuta' in v:
            total_players = 18
        elif 'Vakama' in v:
            total_players = 17
        else:
            raise
    episode = ws.title[1:]
    new_df = extract_tc(ws, cell.row, cell.column)
    new_df['total_players_remaining'] = total_players
    new_df['episode'] = int(re.match('(\d+).*', episode).group(1))
    df = pd.concat([df, new_df], ignore_index=True)
    return df


def append_challenge_data(df, ws, cell, challenge_type):
    search = re.search('F(\d+)', cell.value)
    if not search:
        # We don't have information about the "final" amount, so we don't fill this in
        final = None
    else:
        final = int(search.group(1))

    if challenge_type == 'RC':
        extract_f = extract_rc_challenge
    elif challenge_type == 'IC':
        extract_f = extract_ic_challenge
    else:
        raise ValueError

    episode = ws.title[1:]
    new_df = extract_f(ws, cell.row, cell.column)
    new_df['total_players_remaining'] = final
    new_df['episode'] = int(re.match('(\d+).*', episode).group(1))
    try:
        df = pd.concat([df, new_df], ignore_index=True)
    except:
        import pdb
        pdb.set_trace()
    return df


def append_episode_data(df, ws, cell):
    episode = ws.title[1:]
    new_df = extract_ep(ws, cell.row, cell.column)
    new_df['episode'] = int(re.match('(\d+).*', episode).group(1))
    df = pd.concat([df, new_df], ignore_index=True)
    return df


def append_rc_data(df, ws, cell):
    return append_challenge_data(df, ws, cell, 'RC')


def append_ic_data(df, ws, cell):
    return append_challenge_data(df, ws, cell, 'IC')


def ic_transform(df, full_name_dict_to_id):
    df['win'] = df['win?']
    if 1 in df:
        df['win'] = df['win?'].fillna(df[1])
    if 'win? ' in df:
        df['win'] = df['win'].fillna(df['win? '])

    if 'sitout' in df:
        df['sitout'] = df['SO'].fillna('sitout')
    else:
        df['sitout'] = df['SO']

    merge_key = df['contestant'].str.replace(
        ' ', '_') + '_' + df['season_id'].astype(str)
    df['contestant_id'] = merge_key.map(full_name_dict_to_id)

    rel_columns = df.columns[df.columns.isin(
        ['win?', 'contestant', 'SO', 'win? '])]
    df.rename(columns={'#people': 'team', 'win%': 'win_pct',
                       'total win%': 'episode_win_pct'}, inplace=True)
    df.drop(columns=rel_columns, inplace=True)

    df = df[df['win'].notnull()].reset_index(drop=True)

    return df


def extract_episodal_data(ws):
    results = {'tribal_council': pd.DataFrame(),
               'reward_challenge': pd.DataFrame(),
               'immunity_challenge': pd.DataFrame(),
               'overall_episode': pd.DataFrame()}

    for row in ws['A1:AZ100']:
        for cell in row:
            if isinstance(cell.value, str):

                if 'Tribal Council voting' in cell.value:
                    # we have to do this first to get the total players/ final N
                    results['tribal_council'] = append_tc_data(
                        results['tribal_council'], ws, cell)

                elif ('RC' in cell.value) and (cell.row > 1) and (len(cell.value) < 25):
                    results['reward_challenge'] = append_rc_data(
                        results['reward_challenge'], ws, cell)

                elif (('IC' in cell.value) or ('Immunity challenge' in cell.value)) and (cell.row > 1) and (len(cell.value) < 25):
                    results['immunity_challenge'] = append_ic_data(
                        results['immunity_challenge'], ws, cell)

                elif (cell.row == 1) and (cell.column == 2):
                    results['overall_episode'] = append_episode_data(
                        results['overall_episode'], ws, cell)

    for challenge_key in ['reward_challenge', 'immunity_challenge']:
        if results[challenge_key].shape[0] == 0:
            continue
        if results[challenge_key]['total_players_remaining'].isna().any():
            try:
                results[challenge_key]['total_players_remaining'] = results['tribal_council']['total_players_remaining'].min()
            except KeyError:
                break

            n_occur = results[challenge_key].groupby(
                'contestant')['episode'].transform(lambda x: np.argsort(x.index))

            if challenge_key == 'immunity_challenge':
                results[challenge_key]['total_players_remaining'] -= n_occur
            else:
                results[challenge_key]['challenge_number'] = n_occur + 1
    return results


def find_data_for_season(wb):
    results = {'tribal_council': pd.DataFrame(),
               'reward_challenge': pd.DataFrame(),
               'immunity_challenge': pd.DataFrame(),
               'overall_episode': pd.DataFrame()}

    sheets = [s for s in wb.worksheets if re.match('e\d+', s.title)]

    for sheet in sheets:
        df_dict = extract_episodal_data(sheet)

        for k, v in df_dict.items():
            current_cols = results[k].columns
            if any(x not in v.columns for x in current_cols):
                v = v.reindex(columns=current_cols)
            try:
                results[k] = pd.concat([results[k], v])
            except:
                import pdb
                pdb.set_trace()
    for k, df in results.items():
        if df.shape[0] == 0:
            continue

        if k != 'overall_episode':
            df['tc_number'] = df['total_players_remaining'].max() + 1 - \
                df['total_players_remaining']

            df.sort_values(by='tc_number', inplace=True)
        df.reset_index(drop=True, inplace=True)
    return results


season_type_map = {
    's': 'Survivor',
    'au': 'Australian Survivor',
    'nz': 'Survivor New Zealand',
    'ssa': 'Survivor South Africa'
}


def extract_episode_stats(eng, data_path=None, asof=None):

    if data_path is None:
        data_path = os.path.join(os.path.dirname(__file__),
                                 '../../../data/raw/truedorks')
    new_seasons = search_for_new_seasons(eng, asof=asof)

    new_seasons_df = pd.read_sql(con=eng,
                                 sql='SELECT * FROM survivor.season').set_index('name').loc[new_seasons]

    search_string = new_seasons_df['type'].map(
        {v: k for k, v in season_type_map.items()}) + new_seasons_df['season_number'].astype(int).astype(str)
    print(search_string)

    subset_dl = [y for y in search_string]
    pull_and_save_excels(None, subset=subset_dl, dest_folder=data_path)

    raw_files = glob.glob('{data_path}/*'.format(data_path=data_path))

    raw_files = [f for f in raw_files if any(
        re.search(y, f) for y in search_string)]
    results = {'tribal_council': pd.DataFrame(),
               'reward_challenge': pd.DataFrame(),
               'immunity_challenge': pd.DataFrame(),
               'overall_episode': pd.DataFrame()}
    for f in raw_files:
        if '~' in f or 'total_voting' in f:
            continue
        print('Extracting data from {path}...'.format(
            path=os.path.realpath(f)))
        wb = load_workbook(f, data_only=True)
        season_dfs = find_data_for_season(wb)
        season_raw = os.path.basename(f).split('_')[0]
        season_parts = re.match('(.*?)(\d+)', season_raw).groups(0)
        season_type = season_type_map[season_parts[0]]
        season_num = int(season_parts[1])

        for k, df in season_dfs.items():

            df['season_id'] = get_season_id_by_number_type(
                eng, season_num, season_type)
            if not df.empty:
                df['episode_id'] = df[['episode', 'season_id']].apply(
                    lambda x: get_ep_id_by_number(eng, x[0], x[1]), axis=1)
            try:
                results[k] = pd.concat([results[k], df])
            except:
                import pdb
                pdb.set_trace()
        print('Extraction complete!')
    return results
